import re
from collections import OrderedDict, Counter
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
TARGETS = {
 'ahg-actor-manage':'GLAM/DAM - Authority Records','ahg-information-object-manage':'GLAM/DAM - Archival Description',
 'ahg-dam':'GLAM/DAM - Digital Assets (DAM)','ahg-display':'GLAM/DAM - GLAM Browse & Display',
 'ahg-museum':'GLAM/DAM - Museum Collections','ahg-library':'GLAM/DAM - Library','ahg-gallery':'GLAM/DAM - Gallery',
 'ahg-iiif-collection':'GLAM/DAM - IIIF','ahg-repository-manage':'GLAM/DAM - Repository','ahg-accession-manage':'GLAM/DAM - Accessions',
 'ahg-condition':'AI Condition Assessment','ahg-spectrum':'Condition (Spectrum 5.1)','ahg-provenance-ai':'GLAM/DAM - Provenance',
 'ahg-loan':'GLAM/DAM - Loans','ahg-exhibition':'GLAM/DAM - Exhibitions','ahg-vendor':'Vendor','ahg-donor-manage':'Donor',
 'ahg-heritage-manage':'Heritage Assets','ahg-ipsas':'Heritage Assets (IPSAS)','ahg-reports':'Reports',
 'ahg-research':'Research Services','ahg-rdm':'Research Services (RDM)','ahg-researcher-manage':'Research Admin',
 'ahg-semantic-search':'Knowledge Platform','ahg-discovery':'Knowledge Platform (Discovery)',
 'ahg-access-request':'Access Requests','ahg-security-clearance':'Security & Compliance','ahg-acl':'Security & Compliance (Access Control)',
 'ahg-ai-compliance':'Security & Compliance (AI Act)','ahg-privacy':'Privacy & Data Protection','ahg-cdpa':'Privacy & Data Protection (Register)',
 'ahg-ai-services':'AI Condition Assessment (AI Services)','ahg-workflow':'Condition (Spectrum 5.1 Workflow)',
 'ahg-rights':'Rights & Licensing','ahg-extended-rights':'Rights & Licensing (Extended)','ahg-rights-holder-manage':'Rights & Licensing (Holders)',
 'ahg-marketplace':'Marketplace / Sales & Payouts','ahg-cart':'Marketplace (Cart & Checkout)',
 'ahg-doi-manage':'DOI Management','ahg-doi':'DOI Management (DOIs)','ahg-ric':'Records in Contexts (RiC)',
 'ahg-data-migration':'Data Migration','ahg-ingest':'Data Ingest','ahg-scan':'Data Ingest (Scanning)','ahg-ftp-upload':'Data Ingest (Bulk/FTP)',
 'ahg-backup':'Backup & Maintenance','ahg-dedupe':'Duplicate Detection','ahg-pdf-tools':'TIFF to PDF Merge',
 'ahg-preservation':'Digital Preservation','ahg-ocfl':'Digital Preservation (OCFL)','ahg-archivematica':'Digital Preservation (Archivematica)',
 'ahg-integrity':'Checksums & Integrity','ahg-naz':'Jurisdiction Compliance (NAZ)','ahg-nmmz':'Jurisdiction Compliance (NMMZ)',
 'ahg-narssa':'Jurisdiction Compliance (NARSSA)',
 # NEW:
 'ahg-feedback':'Feedback & Corrections','ahg-favorites':'Favourites & Research Folders',
}
OVR = {'Authority Record Management':('Authority Record - full CRUD','Built'),
 'Description CRUD and Metadata Standards':('Archival Description - full CRUD','Built'),
 'Asset Catalogue':('Digital Asset - full CRUD','Built'),
 'Digital Objects and Media':('Archival Description with digital object','In progress')}
lines=open('docs/help/functions.md').read().split('\n'); area=mod=grp=None; rows=[]
for l in lines:
    m=re.match(r'^## (.+?) \((ahg-[a-z0-9-]+)\)\s*$',l)
    if m: mod=m.group(2); area=TARGETS.get(mod); grp=None; continue
    if l.startswith('## '): area=mod=None; continue
    if area and l.startswith('### '): grp=l[4:].strip(); continue
    if area and grp and l.startswith('- ') and 'CLI Command' not in grp:
        fn=re.sub(r'`([^`]*)`',r'\1',l[2:].strip())
        if fn and not fn.lower().startswith('no user-facing'):
            scen,status = OVR.get(grp,(f"{area} - {grp}",'Suggested'))
            rows.append([area,mod,grp,fn[:400],scen,status])
scen_meta=OrderedDict(); order={'Built':0,'In progress':1,'Planned':2,'Proposed':3,'Suggested':4}
for area,mod,grp,fn,scen,status in rows:
    d=scen_meta.setdefault(scen,{'area':area,'mods':set(),'n':0,'status':status,'groups':set()})
    d['mods'].add(mod); d['n']+=1; d['groups'].add(grp)
    if order[status]<order[d['status']]: d['status']=status
def steps(scen,groups):
    if any(k in scen.lower() for k in ['crud','catalogu','management','register','record']): return 'Browse -> Add/Create -> View -> Edit -> Delete'
    return "Walkthrough: "+" / ".join(sorted(groups))
# manual RiC-view-per-form scenarios (separate wav each)
RIC_VIEWS=[
 ('RiC View - Archival Description','Records in Contexts (RiC)','ahg-ric, ahg-information-object-manage','Open an archival description -> click the RiC button -> scroll to the RiC relationships / knowledge graph','Proposed'),
 ('RiC View - Authority Record','Records in Contexts (RiC)','ahg-ric, ahg-actor-manage','Open an authority record -> click the RiC button -> scroll to the RiC portion','In progress'),
 ('RiC View - Accession','Records in Contexts (RiC)','ahg-ric, ahg-accession-manage','Open an accession -> click the RiC button -> scroll to the RiC portion','Proposed'),
 ('RiC View - Donor','Records in Contexts (RiC)','ahg-ric, ahg-donor-manage','Open a donor record -> click the RiC button -> scroll to the RiC portion','Proposed'),
 ('RiC View - Repository','Records in Contexts (RiC)','ahg-ric, ahg-repository-manage','Open a repository -> click the RiC button -> scroll to the RiC portion','Proposed'),
]
wb=Workbook()
def style(ws,n):
    f=PatternFill('solid',fgColor='1F3A5F')
    for c in range(1,n+1):
        cc=ws.cell(1,c); cc.font=Font(bold=True,color='FFFFFF'); cc.fill=f; cc.alignment=Alignment(vertical='center',wrap_text=True)
    ws.freeze_panes='A2'
sc=wb.active; sc.title='Scenarios'
sc.append(["#","Scenario","Area","Module(s)","# functions","Suggested steps","Status","Output name (.wav/.mp4)","Notes"])
i=0
for scen,d in scen_meta.items():
    i+=1; sc.append([i,scen,d['area'],", ".join(sorted(d['mods'])),d['n'],steps(scen,d['groups']),d['status'],
                     scen if d['status'] in('Built','In progress') else '',''])
for scen,area,mods,st,status in RIC_VIEWS:
    i+=1; sc.append([i,scen,area,mods,'-',st,status,scen,'RiC view - separate wav'])
style(sc,9)
for w,c in zip([4,52,34,32,11,58,12,34,22],range(1,10)): sc.column_dimensions[get_column_letter(c)].width=w
fn=wb.create_sheet('Functions')
fn.append(["Area","Module (package)","Function group","Function","Suggested scenario","Status","Include (Y/N)"])
for r in rows: fn.append(r+[''])
style(fn,7)
for w,c in zip([32,28,32,74,48,12,11],range(1,8)): fn.column_dimensions[get_column_letter(c)].width=w
for rr in range(2,fn.max_row+1): fn.cell(rr,4).alignment=Alignment(wrap_text=True,vertical='top')
out='/usr/share/nginx/heratio-dev/test-results/heratio-demo-scenarios.xlsx'; wb.save(out)
print(f"{len(rows)} functions -> {len(scen_meta)} auto scenarios + {len(RIC_VIEWS)} RiC-view scenarios = {i} rows")
fb=sum(1 for r in rows if r[0].startswith('Feedback')); fav=sum(1 for r in rows if r[0].startswith('Favourites'))
print(f"  Feedback & Corrections: {fb} funcs | Favourites & Research Folders: {fav} funcs | RiC-view scenarios: {len(RIC_VIEWS)}")
